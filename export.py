"""Export attack data to Excel (.xlsx)."""

from io import BytesIO
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1E2433")
HEADER_FONT = Font(bold=True, color="E2E8F0")


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)


def _style_header_row(ws):
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def build_attacks_workbook(db) -> BytesIO:
    wb = Workbook()

    # ── Sheet 1: All attacks ──
    ws = wb.active
    ws.title = "All Attacks"
    columns = [
        "id", "timestamp", "ip", "country", "city",
        "method", "path", "user_agent", "payload", "headers",
    ]
    ws.append([c.replace("_", " ").title() for c in columns])
    _style_header_row(ws)

    for row in db.execute(
        "SELECT * FROM attacks ORDER BY id DESC"
    ).fetchall():
        ws.append([row[c] for c in columns])
    _auto_width(ws)

    # ── Sheet 2: Top IPs ──
    ws_ips = wb.create_sheet("Top IPs")
    ws_ips.append(["IP", "Country", "Hits"])
    _style_header_row(ws_ips)
    for row in db.execute("""
        SELECT ip, country, COUNT(*) as hits
        FROM attacks GROUP BY ip ORDER BY hits DESC LIMIT 100
    """).fetchall():
        ws_ips.append([row["ip"], row["country"], row["hits"]])
    _auto_width(ws_ips)

    # ── Sheet 3: Top Paths ──
    ws_paths = wb.create_sheet("Top Paths")
    ws_paths.append(["Path", "Hits"])
    _style_header_row(ws_paths)
    for row in db.execute("""
        SELECT path, COUNT(*) as hits
        FROM attacks GROUP BY path ORDER BY hits DESC LIMIT 100
    """).fetchall():
        ws_paths.append([row["path"], row["hits"]])
    _auto_width(ws_paths)

    # ── Sheet 4: By Country ──
    ws_countries = wb.create_sheet("By Country")
    ws_countries.append(["Country", "Hits"])
    _style_header_row(ws_countries)
    for row in db.execute("""
        SELECT country, COUNT(*) as hits
        FROM attacks GROUP BY country ORDER BY hits DESC
    """).fetchall():
        ws_countries.append([row["country"], row["hits"]])
    _auto_width(ws_countries)

    # ── Sheet 5: Daily summary ──
    ws_daily = wb.create_sheet("Daily Summary")
    ws_daily.append(["Date", "Hits"])
    _style_header_row(ws_daily)
    for row in db.execute("""
        SELECT DATE(timestamp) as day, COUNT(*) as hits
        FROM attacks GROUP BY day ORDER BY day DESC
    """).fetchall():
        ws_daily.append([row["day"], row["hits"]])
    _auto_width(ws_daily)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_filename() -> str:
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    return f"honeytrap-attacks-{stamp}.xlsx"
